

import tkinter
from tkinter import * # tkinter의 모든 함수 가져오기
from tkinter import messagebox, filedialog
import os
from pathlib import Path
import openpyxl
import os.path
from openpyxl.worksheet.table import Table, TableStyleInfo
import tkinter.ttk
import tkinter as tk

#왼쪽 목록 관련 함수
def del_t(): #오른쪽 트리 삭제용
    tree.delete(*tree.get_children())
def del_t2(): #오른쪽 트리 삭제용
    tree2.delete(*tree2.get_children())

def l_tree1():
    global temp_sheet
    del_t()
    og_sheets_row()
    setlist()

    temp_sheet=og_sheets[0]
    for i in range(1,og_row[0]):
        tree.insert('', 'end', text="", values=og_l[0][i])
    tree.place(x=10,y=200)
def l_tree2():
    global temp_sheet
    del_t()
    og_sheets_row()
    setlist()

    temp_sheet = og_sheets[1]
    for i in range(1, og_row[1]):
        tree.insert('', 'end', text="", values=og_l[1][i])
    tree.place(x=10, y=200)
def l_tree3():
    global temp_sheet
    del_t()
    og_sheets_row()
    setlist()

    temp_sheet = og_sheets[2]
    for i in range(1, og_row[2]):
        tree.insert('', 'end', text="", values=og_l[2][i])
    tree.place(x=10, y=200)
def l_tree4():
    global temp_sheet
    temp_sheet=[]
    del_t()
    og_sheets_row()
    setlist()

    temp_sheet = og_sheets[3]
    for i in range(1, og_row[3]):
        tree.insert('', 'end', text="", values=og_l[3][i])
    print(temp_sheet.cell(2, 1).value)
    temp_sheet.cell(2, 1).value = 3
    print(temp_sheet.cell(2, 1).value)
    print(og_sheets[3].cell(2,1).value)
    tree.place(x=10, y=200)
def l_tree5():
    global temp_sheet
    temp_sheet=[]
    del_t()
    og_sheets_row()
    setlist()
    temp_sheet = og_sheets[4]
    for i in range(1, og_row[4]):
        tree.insert('', 'end', text="", values=og_l[4][i])
    print(temp_sheet.cell(2, 1).value)
    temp_sheet.cell(2, 1).value=3
    print(temp_sheet.cell(2,1).value)
    tree.place(x=10, y=200)



def og_sheets_row(): #왼쪽 시트별 길이 저장 =>og_row(5개 기준)
    count = 0
    for i in range(len(og_sheets)):
        for rows in og_sheets[i].iter_rows():  # ws시트 row 길이를 count에 저장
            count += 1
        og_row[i]=count
        count=0
def setlist(): #셀 값 저장 => og_l 시트 5개 기준(column 3개)
    og_sheets_row() #사용할때마다 row를 다시 구한다
    row=[]
    for i in range(len(og_l)): #og_l리스트 길이만큼(5)
         for k in range(1, (og_row[i] + 1)):  # og_l[i] row 길이만큼 반복
            for j in range(1, 4): #column 1,2,3 저장 (품명,가격,수량)
                if(k==1):
                    row.append(og_sheets[i].cell(k,j).value)
                elif(type(og_sheets[i].cell(k, j).value)==float):
                    row.append(int(og_sheets[i].cell(k,j).value))
                    # print(type(int(og_sheets[i].cell(k, j).value)))
                elif((og_sheets[i].cell(k, j).value==None)):
                    row.append(int(0))
                else:
                    row.append(og_sheets[i].cell(k, j).value)

            og_l[i].append(row)
            # print(og_l[i])
            row = []
    # for i in range(len(og_l)):
    #     print(og_l[i])
def left_double(event): #왼쪽 물품 더블클릭
    def close():
        center_tree()
        count_item.quit()
        count_item.destroy()
    def go(): #확인 버튼
        num=int(amount.get()) #입력된 텍스트(수량)저장
        selectedItem = tree.selection()[0]  # tree 선택한 위치 받기
        #물품명 단가 수량 금액

        row=[] #지역변수 리셋 필요 없음
        if(((tree.item(selectedItem)['values'][2])==None)| (tree.item(selectedItem)['values'][2]<num)):
            messagebox.showinfo("","수량보다 많이 입력하였습니다.")
        else:
            # tree.item(selectedItem)['values'][2]=num-int(tree.item(selectedItem)['values'][2])
            print(tree.item(selectedItem)['values'][2])

            row.append(tree.item(selectedItem)['values'][0]) #물품명
            row.append(tree.item(selectedItem)['values'][1]) #단가
            row.append(num) #수량
            row.append(row[1]*num) #금액
            # messagebox.showinfo("",tree.item(selectedItem)['values'][0]) 물품명만 받기
            new_p.append(row) #new_p에 저장(선택한 값 모두 받기
        print(new_p)
        close()
    def go_enter(event): #엔터 사용을 위한 함수
        num = int(amount.get())
        selectedItem = tree.selection()[0]
        row = []
        if (((tree.item(selectedItem)['values'][2]) == None) | (tree.item(selectedItem)['values'][2] < num)):
            messagebox.showinfo("", "수량보다 많이 입력하였습니다.")
        else:
            row.append(tree.item(selectedItem)['values'][0])  # 물품명
            row.append(tree.item(selectedItem)['values'][1])  # 단가
            row.append(num)  # 수량
            row.append(row[1] * num)  # 금액
            # messagebox.showinfo("",tree.item(selectedItem)['values'][0]) 물품명만 받기
            new_p.append(row)
        print(new_p)
        close()


    count_item = Tk()  # 불러오기 하면 나오는 화면

    count_item.geometry("200x150+500+300")  # 창의 크기
    count_item.title("수량 입력")  # 창의 제목
    count_item.option_add("*Font", "맑은고딕 14")  # 전체 폰트

    ontk = Label(count_item) #수량 레이블
    ontk.config(text="수량", width=10, relief="solid")
    ontk.pack(side="top", pady=10)

    amount = Entry(count_item) #수량 엔트리 go_enter 연결
    amount.config(width=10, relief="solid", borderwidth=0)
    amount.focus()
    amount.bind("<Return>", go_enter)
    amount.place(x=60,y=50)
    amount.pack()

    conf = Button(count_item, text="확인") #확인 버튼
    conf.config(width=10, height=3, command=go) #go 연결
    # conf.place(x=30,y=200)
    conf.pack(side="bottom",pady=10)
    count_item.mainloop()
def center_tree():
    del_t2()
    for i in range(len(new_p)):
        tree2.insert('', 'end', text="", values=new_p[i])
    tree2.place(x=500, y=200)


def trees1():
    del_t()
    row=[]
    b=[]
    for x in range(2, (og_sheets[0].max_row + 1)):
        for y in range(1, 4):
            row.append(og_sheets[0].cell(x, y).value)
        b.append(row)
        row = []
        tree.insert('', 'end', text="", values=b[x - 2])
def trees2():
    del_t()
    row=[]
    b=[]
    for x in range(2, (og_sheets[1].max_row + 1)):
        for y in range(1, 4):
            row.append(og_sheets[1].cell(x, y).value)
        b.append(row)
        row = []
        tree.insert('', 'end', text="", values=b[x - 2])
def trees3():
    del_t()
    row=[]
    b=[]
    for x in range(2, (og_sheets[2].max_row + 1)):
        for y in range(1, 4):
            row.append(og_sheets[2].cell(x, y).value)
        b.append(row)
        row = []
        tree.insert('', 'end', text="", values=b[x - 2])
def trees4():
    del_t()
    row=[]
    b=[]
    for x in range(2, (og_sheets[3].max_row + 1)):
        for y in range(1, 4):
            row.append(og_sheets[3].cell(x, y).value)
        b.append(row)
        row = []
        tree.insert('', 'end', text="", values=b[x - 2])
def trees5():
    del_t()
    row=[]
    b=[]
    for x in range(2,(og_sheets[4].max_row+1)):
        for y in range(1,4):
            row.append(og_sheets[4].cell(x,y).value)
        b.append(row)
        row=[]
        tree.insert('', 'end', text="", values=b[x-2])


if __name__ == "__main__":
#시트기준


#빈소 특 101,102,201,202, 안치1, 안치2, 안치3
    # sp101
    # sp102
    # sp201
    # sp202


    home = 'xl/test.xlsx'
    info_xl='xl/personal.xlsx'

    og_file= openpyxl.load_workbook(home, data_only=True) #초기 시트 위치 저장(값으로)
    info_file=openpyxl.load_workbook(info_xl,data_only=True) #개인정보, 빈소별 물품정보 저장 공간(값으)

    inf_sheets=[info_file['빈소1']]
    og_sheets=[og_file['식당판매'], og_file['매점판매'], og_file['장의용품'], og_file['상복'], og_file['기타']]  #시트 리스트에 저장 시트 이름 바꾸면 같이 바꿔야 함
    og_row=['','','','',''] #길이 저장
    og_l=[[],[],[],[],[]] #column 2개에 있는 cell info each list에 저장
    new_l=[] #불러오거나 저장핳때 사용할 예정
    global temp_l #목록만 기록
    temp_l=[]
    temp_sheet=[]

    global og_p #왼쪽 목록 폼 출력용
    global new_p #중앙 목록 폼 출력용
    new_p=[]
    global count




    win = tk.Tk() # 창 생성
    win.geometry("1200x720") # 창의 크기
    win.title("장례식장 재고관리 프로그램 Ver1.221123") # 창의 제목
    win.option_add("*Font", "맑은고딕 12") # 전체 폰트

    #-------------------------------------------------

    tree = tkinter.ttk.Treeview(win, columns=["one", "two", "three"],
                                displaycolumns=["one", "two", "three"], height=24)  # 3개 창 생성

    tree.column("#0", width=10, anchor="center")  # 1
    tree.heading("#0", text="", anchor="center")

    tree.column("#1", width=90, anchor="center")  # 2
    tree.heading("#1", text="물품명", anchor="center")

    tree.column("#2", width=100, anchor="center")  # 3
    tree.heading("#2", text="단가", anchor="center")

    tree.column("#3", width=100, anchor="center")  # 4
    tree.heading("#3", text="수량", anchor="center")

    #-------------------------------------------------

    tree2 = tkinter.ttk.Treeview(win, columns=["one", "two", "three","four"],
                                displaycolumns=["one", "two", "three","four"], height=24)  # 4개 창 생성

    tree2.column("#0", width=10, anchor="center")  # 0
    tree2.heading("#0", text="", anchor="center")

    tree2.column("#1", width=90, anchor="center")  # 1
    tree2.heading("#1", text="물품명", anchor="center")

    tree2.column("#2", width=100, anchor="center")  # 2
    tree2.heading("#2", text="단가", anchor="center")

    tree2.column("#3", width=100, anchor="center")  # 3
    tree2.heading("#3", text="수량", anchor="center")

    tree2.column("#4", width=100, anchor="center")  # 4
    tree2.heading("#4", text="금액", anchor="center")


    #-------------------------------------------------

    시트1 = Button(win, text = "식당판매")
    시트1.config(width=7,height=2,command=trees1)
    시트1.place(x=10,y=10)

    시트2 = Button(win, text = "매점판매")
    시트2.config(width=7,height=2,command=trees2)
    시트2.place(x=100,y=10)

    시트3 = Button(win, text = "장의용품")
    시트3.config(width=7,height=2,command=trees3)
    시트3.place(x=190,y=10)

    시트4 = Button(win, text = "상복")
    시트4.config(width=7,height=2,command=trees4)
    시트4.place(x=280,y=10)

    시트5 = Button(win, text = "기타")
    시트5.config(width=7,height=2,command=trees5)
    시트5.place(x=370,y=10)

    tree.place(x=10,y=200)
    tree.bind("<Double-Button-1>",left_double)
    tree2.place(x=500,y=200)
    # tree2.bind("<Double-Button-1>",center_double)



    win.mainloop() # 창 실행